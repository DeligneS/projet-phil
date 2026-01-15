"""Excel document parser using openpyxl."""

import io

from openpyxl import load_workbook


def parse_excel(content: bytes) -> str:
    """Extract text content from an Excel file.

    Args:
        content: Raw bytes of the Excel file.

    Returns:
        Extracted text from all sheets, formatted as tables.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_text = [f"=== Sheet: {sheet_name} ==="]

        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                if cell.value is not None:
                    row_values.append(str(cell.value))
            if row_values:
                sheet_text.append(" | ".join(row_values))

        if len(sheet_text) > 1:  # Has content beyond the header
            text_parts.append("\n".join(sheet_text))

    wb.close()
    return "\n\n".join(text_parts)
