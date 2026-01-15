"""Excel export for evaluation results."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.evaluation.llm_evaluator import EvaluationResult


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a string to be used as an Excel sheet name.

    Excel sheet names have restrictions:
    - Max 31 characters
    - Cannot contain: [ ] : * ? / \\

    Args:
        name: The desired sheet name.

    Returns:
        Sanitized sheet name.
    """
    # Remove invalid characters
    invalid_chars = ["[", "]", ":", "*", "?", "/", "\\"]
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, "_")

    # Truncate to 31 characters
    return sanitized[:31]


def create_excel_report(evaluations: list[EvaluationResult]) -> io.BytesIO:
    """Create an Excel report with one sheet per student.

    Args:
        evaluations: List of evaluation results for all students.

    Returns:
        BytesIO containing the Excel file.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for evaluation in evaluations:
        sheet_name = sanitize_sheet_name(evaluation.student_name)
        ws = wb.create_sheet(title=sheet_name)

        current_row = 1

        # Title
        ws.cell(row=current_row, column=1, value=f"Évaluation - {evaluation.student_name}")
        ws.cell(row=current_row, column=1).font = title_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 2

        # Criteria table header
        headers = ["Critère", "Note", "Note Max", "Commentaire"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Criteria rows
        for criterion in evaluation.criteres:
            ws.cell(row=current_row, column=1, value=criterion.nom).border = border
            ws.cell(row=current_row, column=2, value=criterion.note).border = border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value=criterion.note_max).border = border
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=criterion.commentaire).border = border
            ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True)
            current_row += 1

        current_row += 1

        # Final grade
        ws.cell(row=current_row, column=1, value="Note finale").font = header_font
        ws.cell(
            row=current_row, column=2, value=f"{evaluation.note_finale} / {evaluation.note_max}"
        )
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=14)
        current_row += 2

        # General feedback section
        ws.cell(row=current_row, column=1, value="Feedback général").font = header_font
        current_row += 1

        feedback_cell = ws.cell(row=current_row, column=1, value=evaluation.feedback_general)
        feedback_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 5, end_column=4)
        current_row += 7

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 50

    # Create BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def create_summary_sheet(wb: Workbook, evaluations: list[EvaluationResult]) -> None:
    """Add a summary sheet with all students' grades.

    Args:
        wb: The workbook to add the summary to.
        evaluations: List of all evaluation results.
    """
    ws = wb.create_sheet(title="Résumé", index=0)

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Étudiant", "Note finale", "Note max"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, evaluation in enumerate(evaluations, start=2):
        ws.cell(row=row, column=1, value=evaluation.student_name).border = border
        ws.cell(row=row, column=2, value=evaluation.note_finale).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=evaluation.note_max).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def create_excel_report_with_summary(evaluations: list[EvaluationResult]) -> io.BytesIO:
    """Create an Excel report with a summary sheet and one sheet per student.

    Args:
        evaluations: List of evaluation results for all students.

    Returns:
        BytesIO containing the Excel file.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for evaluation in evaluations:
        sheet_name = sanitize_sheet_name(evaluation.student_name)
        ws = wb.create_sheet(title=sheet_name)

        current_row = 1

        # Title
        ws.cell(row=current_row, column=1, value=f"Évaluation - {evaluation.student_name}")
        ws.cell(row=current_row, column=1).font = title_font
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 2

        # Criteria table header
        headers = ["Critère", "Note", "Note Max", "Commentaire"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Criteria rows
        for criterion in evaluation.criteres:
            ws.cell(row=current_row, column=1, value=criterion.nom).border = border
            ws.cell(row=current_row, column=2, value=criterion.note).border = border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value=criterion.note_max).border = border
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=criterion.commentaire).border = border
            ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True)
            current_row += 1

        current_row += 1

        # Final grade
        ws.cell(row=current_row, column=1, value="Note finale").font = header_font
        ws.cell(
            row=current_row, column=2, value=f"{evaluation.note_finale} / {evaluation.note_max}"
        )
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=14)
        current_row += 2

        # General feedback section
        ws.cell(row=current_row, column=1, value="Feedback général").font = header_font
        current_row += 1

        feedback_cell = ws.cell(row=current_row, column=1, value=evaluation.feedback_general)
        feedback_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 5, end_column=4)
        current_row += 7

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 50

    # Add summary sheet at the beginning
    create_summary_sheet(wb, evaluations)

    # Create BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
